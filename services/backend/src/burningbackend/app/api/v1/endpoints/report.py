from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from starlette.responses import StreamingResponse

from burningbackend.app.models.history import History
from burningbackend.app.models.movie import Movie

router = APIRouter()


def _summarize_products(orders: list[History]) -> dict:
    summary: dict = {}
    for order in orders:
        for product in order.products:
            entry = summary.setdefault(product.name, {"amount": 0, "total": 0.0, "price": product.price})
            entry["amount"] += product.amount
            entry["total"] += product.price * product.amount
    return summary


def _get(d: dict, key: str, field: str) -> float | int:
    return d[key][field] if key in d else 0


@router.get("/report", response_description="Excel report retrieved")
async def get_report(
    movie: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> StreamingResponse:
    if not movie and not (start_date or end_date):
        raise HTTPException(status_code=422, detail="Provide 'movie' or a date range (start_date / end_date)")

    query_base: dict = {"cancellation": False}
    if movie:
        query_base["movie"] = movie
    if start_date:
        query_base.setdefault("timestamp", {})["$gte"] = start_date
    if end_date:
        query_base.setdefault("timestamp", {})["$lte"] = end_date

    query_team = {**query_base, "isteam": True}
    query_public = {**query_base, "isteam": False}

    data_sold = await History.find(query_public).to_list()
    team_data = await History.find(query_team).to_list()

    total_sold = sum(h.total for h in data_sold)
    total_sold_team = sum(h.total for h in team_data)

    products_dict = _summarize_products(data_sold)
    team_products_dict = _summarize_products(team_data)

    selected_movie = await Movie.find_one({"name": movie}) if movie else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"
    ws2 = wb.create_sheet("Products")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    bold = Font(name="Calibri", bold=True)
    normal = Font(name="Calibri")

    if selected_movie:
        ws.cell(row=1, column=1, value="Movie:").font = bold
        ws.cell(row=1, column=2, value=selected_movie.name).font = normal
        ws.cell(row=1, column=3, value=str(selected_movie.datetime)).font = normal
        ws.cell(row=1, column=4, value=selected_movie.room).font = normal
    else:
        ws.cell(row=1, column=1, value="Period:").font = bold
        ws.cell(row=1, column=2, value=str(start_date or "")).font = normal
        ws.cell(row=1, column=3, value="–").font = normal
        ws.cell(row=1, column=4, value=str(end_date or "")).font = normal

    ws.cell(row=3, column=1, value="Calc:").font = bold
    ws.cell(row=4, column=1, value="Products sold:").font = normal
    ws.cell(row=4, column=2, value=total_sold - _get(products_dict, "Ticket", "total") - _get(products_dict, "Clubkarte", "total") - _get(products_dict, "Pfand", "total")).font = normal
    ws.cell(row=5, column=1, value="Tickets sold:").font = normal
    ws.cell(row=5, column=2, value=_get(products_dict, "Ticket", "total")).font = normal
    ws.cell(row=6, column=1, value="Clubkarten sold:").font = normal
    ws.cell(row=6, column=2, value=_get(products_dict, "Clubkarte", "total")).font = normal
    ws.cell(row=7, column=1, value="Pfand sold:").font = normal
    ws.cell(row=7, column=2, value=_get(products_dict, "Pfand", "total")).font = normal
    ws.cell(row=8, column=1, value="Total sold:").font = normal
    ws.cell(row=8, column=2, value=total_sold).font = normal
    ws.cell(row=10, column=1, value="Pfand Rück:").font = normal
    ws.cell(row=10, column=2, value=_get(products_dict, "Pfand Rück", "total")).font = normal
    ws.cell(row=11, column=1, value="Total:").font = normal
    ws.cell(row=11, column=2, value=total_sold + _get(products_dict, "Pfand Rück", "total")).font = normal

    ws.cell(row=3, column=4, value="Calc team").font = bold
    ws.cell(row=4, column=4, value="Products team:").font = normal
    ws.cell(row=4, column=5, value=total_sold_team).font = normal
    ws.cell(row=5, column=4, value="Total team:").font = normal
    ws.cell(row=5, column=5, value=total_sold_team).font = normal

    ws.cell(row=12, column=1, value="Tickets:").font = bold
    ws.cell(row=13, column=1, value="Tickets amount:").font = normal
    ws.cell(row=13, column=2, value=_get(products_dict, "Ticket", "amount")).font = normal
    ws.cell(row=14, column=1, value="Freitickets amount:").font = normal
    ws.cell(row=14, column=2, value=_get(products_dict, "Freiticket", "amount")).font = normal
    ws.cell(row=15, column=1, value="Total amount:").font = normal
    ws.cell(row=15, column=2, value=_get(products_dict, "Ticket", "amount") + _get(products_dict, "Freiticket", "amount")).font = normal

    ws2.append(["Name", "Amount_Team", "Team", "Total_Team", "Amount", "Price", "Total"])
    for name, details in products_dict.items():
        t_amount = team_products_dict[name]["amount"] if name in team_products_dict else 0
        t_price = team_products_dict[name]["price"] if name in team_products_dict else 0
        ws2.append([
            name,
            f"{t_amount} x",
            f"{t_price} €",
            f"{t_price * t_amount} €",
            f"{details['amount']} x",
            f"{details['price']} €",
            f"{details['total']} €",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = movie if movie else "report"
    response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}.xlsx"
    return response
